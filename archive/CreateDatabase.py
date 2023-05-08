import openpyxl
import mariadb
import sys

workbook = openpyxl.load_workbook('companiesofboise2.xlsx')

worksheet = workbook["Companies"]

companies = []
db_companies = []
data = []
worksheet_array = []
technologies = ['frontend_lang', 'frontend_frame', 'backend_lang', 'backend_frame', 'mobile', 'network', 'security',
                'devops', 'analytics', 'database', 'cloud', 'communication']
tech_answers = []
db_technologies = []
db_cmp_tech = []
company_tech = []

for row in worksheet.iter_rows(values_only=True):
    if row[1] is not None:
        companies.append(
            {'name': row[0], 'size': row[1], 'type': row[2], 'street': row[3], 'state': row[4], 'zip': row[5],
             'hire2yr': 1 if row[6] is True else 0, 'intern': 1 if row[7] is True else 0, 'phone': row[8]})

companies.pop(0)

try:
    conn = mariadb.connect(
        user="admin",
        password="1234",
        host="127.0.0.1",
        port=3306,
        database="SWDV_DATABASE"

    )
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(1)

cur = conn.cursor()

try:
    cur.execute("SELECT * FROM company")
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(2)
for name in cur:
    db_companies.append(name)

for company in companies:
    if str(company['name']) not in str(db_companies):
        try:
            cur.execute(
                f'''CALL EnterCompanies("{company["name"]}", "{company["street"]}", "{company["state"]}", "{company["zip"]}", "{company["size"]}", "{company["type"]}", "{company["hire2yr"]}", "{company["intern"]}"); ''')
            print("New Company " + company['name'])
        except mariadb.Error as e:
            print(f"Error connecting to MariaDB Platform: {e}")
            sys.exit(2)

sheet_names = workbook.sheetnames

try:
    cur.execute("SELECT * FROM technologies")
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(2)
for name in cur:
    db_technologies.append(name)

for name in sheet_names:
    for company in db_companies:
        if company[1] == name:
            for row in workbook[name].iter_rows(values_only=True):
                if str(row[0]) in str(technologies):
                    tech_answers.append({'company_id': company[0], 'tech_area': row[0], 'tech_name': row[1],
                                         'usedNow': 1 if row[2] is True else 0, 'shouldTeach': 1 if row[3] is True else 0,
                                         'topThree': 1 if row[4] is True else 0, 'continue': 1 if row[5] is True else 0, 'dateCollected': row[6]})

for answer in tech_answers:
    if str(answer['tech_name']) not in str(db_technologies):
        try:
            cur.execute(
                f'''CALL EnterTechnologies("{answer["tech_name"]}", "{answer["tech_area"]}"); ''')
            print("New Technology " + answer['tech_name'])
        except mariadb.Error as e:
            print(f"Error connecting to MariaDB Platform: {e}")
            sys.exit(2)

try:
    cur.execute("SELECT * FROM technologies")
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(2)
for name in cur:
    db_technologies.append(name)

try:
    cur.execute("SELECT * FROM company_tech")
except mariadb.Error as e:
    print(f"Error connecting to MariaDB Platform: {e}")
    sys.exit(2)
for name in cur:
    db_cmp_tech.append(name)

for answer in tech_answers:
    try:
        cur.execute(f'''
        SELECT technologies.tech_name, technologies.tech_area
        FROM company_tech
        INNER JOIN company ON company.company_id = company_tech.company_id
        INNER JOIN technologies ON technologies.tech_id = company_tech.ct_techid
        WHERE company.company_id = '{answer["company_id"]}'
        ''')
    except mariadb.Error as e:
        print(f"Error connecting to MariaDB Platform: {e}")
        sys.exit(2)
    for name in cur:
        company_tech.append(name)
    if str(answer['tech_name']) in str(db_technologies) and str(answer['tech_name']):
        for technology in db_technologies:
            if str(technology[2]) == str(answer['tech_name']) and str(answer['tech_name']) not in str(company_tech):
                try:
                    cur.execute(
                        f'''CALL EnterCompanyTech("{answer["company_id"]}", "{technology[0]}", "{answer["usedNow"]}", "{answer["shouldTeach"]}", "{answer["topThree"]}", "{answer["continue"]}", "{answer["dateCollected"]}"); ''')
                    print("New Company Tech" + answer['tech_name'])
                except mariadb.Error as e:
                    print(f"Error connecting to MariaDB Platform: {e}")
                    sys.exit(2)
                break

conn.commit()

conn.close()
